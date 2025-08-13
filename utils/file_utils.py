# utils/file_utils.py

import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime
import sqlite3

from utils.config_loader import CONFIG
from utils.date_utils import format_date_for_display

def export_agents_to_excel(main_window, db_manager):
    """Exporte la liste complète des agents vers un fichier Excel."""
    main_window.config(cursor="watch")
    main_window.update_idletasks()
    main_window.set_status("Exportation des agents en cours...")
    
    try:
        agents = db_manager.get_agents()
        if not agents:
            messagebox.showinfo("Information", "Aucun agent à exporter.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            title="Exporter la liste des agents",
            initialfile=f"Export_Agents_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        if not filename: return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agents"
        
        # Correction pour exporter correctement les données de l'objet Agent
        headers = ["ID", "Nom", "Prénom", "PPR", "Grade", "Solde"]
        ws.append(headers)
        
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for agent in agents:
            ws.append([agent.id, agent.nom, agent.prenom, agent.ppr, agent.grade, agent.solde])

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        
        wb.save(filename)
        messagebox.showinfo("Succès", f"Liste des agents exportée avec succès vers\n{filename}")
    except Exception as e:
        messagebox.showerror("Erreur d'écriture", f"Impossible de sauvegarder le fichier : {e}")
    finally:
        main_window.config(cursor="")
        main_window.set_status("Prêt.")

def export_all_conges_to_excel(main_window, db_manager):
    """Exporte la liste complète de tous les congés vers un fichier Excel."""
    main_window.config(cursor="watch")
    main_window.update_idletasks()
    main_window.set_status("Exportation totale en cours...")
    
    try:
        all_conges = db_manager.get_conges()
        if not all_conges:
            messagebox.showinfo("Information", "Aucun congé à exporter.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            title="Exporter tous les congés",
            initialfile=f"Export_Conges_Total_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        if not filename: return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tous les Congés"
        headers = ["Nom Agent", "Prénom Agent", "PPR Agent", "Type Congé", "Début", "Fin", "Jours Pris", "Statut", "Justification", "Intérimaire"]
        ws.append(headers)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        all_agents = {agent.id: agent for agent in db_manager.get_agents()}

        for conge in all_conges:
            agent = all_agents.get(conge.agent_id)
            if not agent:
                agent_nom, agent_prenom, agent_ppr = "Agent", "Supprimé", ""
            else:
                agent_nom, agent_prenom, agent_ppr = agent.nom, agent.prenom, agent.ppr

            interim_info = ""
            if conge.interim_id:
                interim = all_agents.get(conge.interim_id)
                interim_info = f"{interim.nom} {interim.prenom}" if interim else "Agent Supprimé"
            
            row_data = [
                agent_nom, agent_prenom, agent_ppr,
                conge.type_conge, 
                format_date_for_display(conge.date_debut), 
                format_date_for_display(conge.date_fin), 
                conge.jours_pris,
                conge.statut,
                conge.justif or "", 
                interim_info
            ]
            ws.append(row_data)

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
            
        wb.save(filename)
        messagebox.showinfo("Succès", f"Tous les congés ont été exportés avec succès vers\n{filename}")
    except Exception as e:
        messagebox.showerror("Erreur d'écriture", f"Impossible de sauvegarder le fichier : {e}")
    finally:
        main_window.config(cursor="")
        main_window.set_status("Prêt.")

def import_agents_from_excel(main_window, db_manager):
    """Importe des agents depuis un fichier Excel, en ajoutant les nouveaux et mettant à jour les existants."""
    filename = filedialog.askopenfilename(
        title="Sélectionner un fichier Excel à importer",
        filetypes=[("Fichiers Excel", "*.xlsx")]
    )
    if not filename:
        return

    main_window.config(cursor="watch")
    main_window.update_idletasks()
    main_window.set_status("Importation en cours...")
    
    errors = []
    agent_import_headers = CONFIG['agent_import_headers']
    grades = CONFIG['ui']['grades']
    
    # Définition des valeurs par défaut
    default_grade = grades[0] if grades else "Administrateur"
    default_solde = 22.0

    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        header = [str(cell.value or '').lower().strip() for cell in ws[1]]
        
        if not all(h in header for h in agent_import_headers):
            raise ValueError(f"Colonnes requises dans le fichier Excel : {', '.join(agent_import_headers)}")

        col_map = {name: i for i, name in enumerate(header)}
        added_count, updated_count, error_count = 0, 0, 0
        
        db_manager.conn.execute('BEGIN TRANSACTION')
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if all(c is None for c in row): continue

                # 1. Lecture des champs obligatoires
                nom = str(row[col_map['nom']] or '').strip()
                prenom = str(row[col_map['prenom']] or '').strip()
                if not nom or not prenom:
                    raise ValueError("Le nom et le prénom sont obligatoires.")

                # 2. Lecture des champs optionnels avec valeurs par défaut
                ppr = str(row[col_map['ppr']] or '').strip()
                grade = str(row[col_map['grade']] or '').strip()
                solde_str = str(row[col_map['solde']] or '').strip().replace(",", ".")

                # Générer un PPR unique si manquant
                if not ppr:
                    timestamp = datetime.now().strftime('%H%M%S%f')
                    ppr = f"{nom.upper()[:4]}_{prenom.upper()[:4]}_{timestamp}"
                
                # Appliquer le grade par défaut si manquant, sinon le valider
                if not grade:
                    grade = default_grade
                elif grade not in grades:
                    raise ValueError(f"Grade '{grade}' invalide. Grades valides: {', '.join(grades)}")
                
                # Appliquer le solde par défaut si manquant, sinon le valider
                if not solde_str:
                    solde = default_solde
                else:
                    solde = float(solde_str)
                    if solde < 0:
                        raise ValueError(f"Le solde '{solde}' ne peut être négatif.")

                # Recherche et ajout/modification
                agent = db_manager.get_agent_by_ppr(ppr)
                if agent:
                    if not db_manager.modifier_agent(agent.id, nom, prenom, ppr, grade, solde):
                        raise Exception(f"Erreur de mise à jour pour PPR {ppr}.")
                    updated_count += 1
                else:
                    if not db_manager.ajouter_agent(nom, prenom, ppr, grade, solde):
                        raise Exception(f"Le PPR '{ppr}' existe déjà ou une autre erreur est survenue.")
                    added_count += 1

            except (ValueError, TypeError, IndexError) as ve:
                errors.append(f"Ligne {i}: {ve}"); error_count += 1
            except Exception as e:
                errors.append(f"Ligne {i}: Erreur - {e}"); error_count += 1
        
        if error_count > 0:
            db_manager.conn.rollback()
            summary = "Échec de l'importation: Des erreurs ont été détectées.\nL'importation est annulée.\n\nAucune modification n'a été enregistrée."
            if errors:
                summary += "\n\nDétail des erreurs (premières 5):\n" + "\n".join(errors[:5])
            messagebox.showerror("Rapport d'importation", summary)
        else:
            db_manager.conn.commit()
            summary = f"Importation réussie !\n\n- Agents ajoutés : {added_count}\n- Agents mis à jour : {updated_count}"
            messagebox.showinfo("Rapport d'importation", summary)

    except Exception as e:
        db_manager.conn.rollback()
        summary = f"Échec de l'importation: {e}\n\nAucune modification n'a été enregistrée."
        messagebox.showerror("Rapport d'importation", summary)
    finally:
        main_window.config(cursor="")
        main_window.set_status("Prêt.")
        main_window.refresh_all()